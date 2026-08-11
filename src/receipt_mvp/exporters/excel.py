from __future__ import annotations

from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from receipt_mvp.models import ReceiptRecord, Transaction, ValidationStatus


HEADER_FILL = PatternFill("solid", fgColor="0B7285")
HEADER_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(name="맑은 고딕", color="263238", size=9)
SUBTLE_BORDER = Border(bottom=Side(style="thin", color="D9E2E7"))
ERROR_FILL = PatternFill("solid", fgColor="FFE3E3")
WARNING_FILL = PatternFill("solid", fgColor="FFF3BF")
OK_FILL = PatternFill("solid", fgColor="D3F9D8")


def _first_item(transaction: Transaction) -> str | None:
    return transaction.items[0].name if transaction.items else None


def _party_value(transaction: Transaction, role: str, field: str) -> Any:
    party = getattr(transaction, role)
    return getattr(party, field) if party else None


DETAIL_COLUMNS: dict[str, tuple[str, Callable[[ReceiptRecord, Transaction], Any]]] = {
    "transaction_id": ("거래 ID", lambda record, tx: tx.transaction_id),
    "file": ("원본파일", lambda record, tx: record.document.source_file_name),
    "page": ("원본페이지", lambda record, tx: record.document.page_number),
    "platform": ("플랫폼", lambda record, tx: record.document.platform.value),
    "document_type": ("문서유형", lambda record, tx: record.document.document_type),
    "extraction_method": ("추출방식", lambda record, tx: record.document.extraction_method.value),
    "occurred_at": ("거래일시", lambda record, tx: tx.occurred_at),
    "order_number": ("주문번호", lambda record, tx: tx.order_number),
    "approval_number": ("승인번호", lambda record, tx: tx.approval_number),
    "evidence_type": ("증빙유형", lambda record, tx: tx.evidence_type),
    "tax_category": ("과세구분", lambda record, tx: tx.tax_category),
    "card_issuer": ("카드사", lambda record, tx: tx.payment.card_issuer),
    "masked_card_number": ("마스킹 카드번호", lambda record, tx: tx.payment.masked_card_number),
    "installment": ("할부정보", lambda record, tx: tx.payment.installment),
    "payment_method": ("결제수단", lambda record, tx: tx.payment.payment_method),
    "item_name": ("상품명", lambda record, tx: _first_item(tx)),
    "item_quantity": (
        "수량",
        lambda record, tx: tx.items[0].quantity if tx.items else None,
    ),
    "item_unit_price": (
        "단가",
        lambda record, tx: tx.items[0].unit_price if tx.items else None,
    ),
    "item_amount": (
        "상품금액",
        lambda record, tx: tx.items[0].amount if tx.items else None,
    ),
    "seller_name": ("판매자상호", lambda record, tx: _party_value(tx, "seller", "name")),
    "seller_representative": (
        "판매자대표자",
        lambda record, tx: _party_value(tx, "seller", "representative_name"),
    ),
    "seller_business_number": (
        "판매자사업자등록번호",
        lambda record, tx: _party_value(tx, "seller", "business_registration_number"),
    ),
    "seller_phone": ("판매자전화번호", lambda record, tx: _party_value(tx, "seller", "phone_number")),
    "seller_address": ("판매자주소", lambda record, tx: _party_value(tx, "seller", "address")),
    "merchant_name": ("가맹점상호", lambda record, tx: _party_value(tx, "merchant", "name")),
    "merchant_representative": (
        "가맹점대표자",
        lambda record, tx: _party_value(tx, "merchant", "representative_name"),
    ),
    "merchant_number": (
        "가맹점번호",
        lambda record, tx: _party_value(tx, "merchant", "merchant_number"),
    ),
    "merchant_business_number": (
        "가맹점사업자등록번호",
        lambda record, tx: _party_value(tx, "merchant", "business_registration_number"),
    ),
    "merchant_address": ("가맹점주소", lambda record, tx: _party_value(tx, "merchant", "address")),
    "taxable_amount_raw": ("원본과세금액", lambda record, tx: tx.amounts.taxable_amount_raw),
    "supply_amount": ("공급가액", lambda record, tx: tx.amounts.supply_amount),
    "tax_exempt_amount": ("면세금액", lambda record, tx: tx.amounts.tax_exempt_amount),
    "vat_amount": ("부가세액", lambda record, tx: tx.amounts.vat_amount),
    "service_charge": ("봉사료", lambda record, tx: tx.amounts.service_charge),
    "approved_amount": ("승인금액", lambda record, tx: tx.amounts.approved_amount),
    "total_amount": ("합계금액", lambda record, tx: tx.amounts.total_amount),
    "confidence": ("신뢰도", lambda record, tx: tx.confidence),
    "status": ("검증상태", lambda record, tx: tx.validation_status.value),
    "user_modified": ("사용자수정", lambda record, tx: tx.user_modified),
}

DEFAULT_DETAIL_KEYS = [
    "transaction_id",
    "file",
    "page",
    "platform",
    "occurred_at",
    "seller_name",
    "seller_business_number",
    "item_name",
    "supply_amount",
    "tax_exempt_amount",
    "vat_amount",
    "service_charge",
    "total_amount",
    "status",
]

TAX_HEADERS = [
    "거래일자",
    "판매자상호",
    "판매자사업자등록번호",
    "공급가액",
    "부가세액",
    "면세금액",
    "합계금액",
    "증빙유형",
    "원본파일",
    "원본페이지",
    "검증상태",
]


class ExcelExporter:
    def export(
        self,
        records: list[ReceiptRecord],
        output_path: str | Path,
        selected_columns: list[str] | None = None,
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        tax_sheet = workbook.create_sheet("세무정리")
        detail_sheet = workbook.create_sheet("상세내역")
        error_sheet = workbook.create_sheet("오류내역")
        mapping_sheet = workbook.create_sheet("원본매핑")

        self._write_tax_sheet(tax_sheet, records)
        keys = self._detail_keys(selected_columns)
        self._write_detail_sheet(detail_sheet, records, keys)
        self._write_error_sheet(error_sheet, records)
        self._write_mapping_sheet(mapping_sheet, records)
        for sheet in workbook.worksheets:
            self._finish_sheet(sheet)
        workbook.save(path)
        self._verify_reopen(path)
        return path

    def _write_tax_sheet(self, sheet, records: list[ReceiptRecord]) -> None:
        sheet.append(TAX_HEADERS)
        for record, transaction in self._transactions(records):
            sheet.append(
                [
                    transaction.occurred_at.date() if transaction.occurred_at else None,
                    _party_value(transaction, "seller", "name"),
                    _party_value(transaction, "seller", "business_registration_number"),
                    transaction.amounts.supply_amount,
                    transaction.amounts.vat_amount,
                    transaction.amounts.tax_exempt_amount,
                    transaction.amounts.total_amount,
                    transaction.evidence_type,
                    record.document.source_file_name,
                    record.document.page_number,
                    self._status_label(transaction.validation_status),
                ]
            )
        for cell in sheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd"
        for column in ("D", "E", "F", "G"):
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0"
        self._status_conditional_format(sheet, "K", max(2, sheet.max_row))

    def _write_detail_sheet(
        self,
        sheet,
        records: list[ReceiptRecord],
        keys: list[str],
    ) -> None:
        sheet.append([DETAIL_COLUMNS[key][0] for key in keys])
        for record, transaction in self._transactions(records):
            item_count = max(1, len(transaction.items))
            for item_index in range(item_count):
                values = []
                for key in keys:
                    if key in {"item_name", "item_quantity", "item_unit_price", "item_amount"}:
                        item = transaction.items[item_index] if transaction.items else None
                        item_field = {
                            "item_name": "name",
                            "item_quantity": "quantity",
                            "item_unit_price": "unit_price",
                            "item_amount": "amount",
                        }[key]
                        values.append(getattr(item, item_field) if item else None)
                    else:
                        values.append(DETAIL_COLUMNS[key][1](record, transaction))
                sheet.append(values)
        for column_index, key in enumerate(keys, start=1):
            letter = get_column_letter(column_index)
            if key == "occurred_at":
                for cell in sheet[letter][1:]:
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif key in {
                "item_unit_price",
                "item_amount",
                "taxable_amount_raw",
                "supply_amount",
                "tax_exempt_amount",
                "vat_amount",
                "service_charge",
                "approved_amount",
                "total_amount",
            }:
                for cell in sheet[letter][1:]:
                    cell.number_format = "#,##0"
            elif key == "confidence":
                for cell in sheet[letter][1:]:
                    cell.number_format = "0%"
            elif key in {
                "order_number",
                "approval_number",
                "seller_business_number",
                "merchant_business_number",
                "merchant_number",
                "masked_card_number",
            }:
                for cell in sheet[letter][1:]:
                    cell.number_format = "@"
            elif key == "status":
                self._status_conditional_format(sheet, letter, max(2, sheet.max_row))

    def _write_error_sheet(self, sheet, records: list[ReceiptRecord]) -> None:
        sheet.append(
            ["원본파일", "페이지", "오류항목", "인식값", "오류코드", "오류사유", "신뢰도", "사용자수정"]
        )
        for record in records:
            for issue in record.validation_issues:
                evidence = next(
                    (
                        item
                        for item in record.field_evidence
                        if item.standard_path == issue.field_path
                    ),
                    None,
                )
                sheet.append(
                    [
                        record.document.source_file_name,
                        record.document.page_number,
                        issue.field_path,
                        evidence.raw_value if evidence else None,
                        issue.code,
                        issue.message,
                        issue.confidence,
                        issue.user_modified,
                    ]
                )
        for cell in sheet["G"][1:]:
            cell.number_format = "0%"
        if sheet.max_row >= 2:
            sheet.conditional_formatting.add(
                f"A2:H{sheet.max_row}",
                FormulaRule(formula=["$E2<>\"\""], fill=ERROR_FILL),
            )

    def _write_mapping_sheet(self, sheet, records: list[ReceiptRecord]) -> None:
        sheet.append(
            [
                "거래 ID",
                "원본파일",
                "페이지",
                "원본 항목명",
                "표준 항목명",
                "원본 값",
                "변환된 값",
                "파서",
                "추출 방식",
                "변환 규칙",
                "신뢰도",
                "사용자 수정",
            ]
        )
        for record in records:
            for evidence in record.field_evidence:
                sheet.append(
                    [
                        evidence.transaction_id,
                        record.document.source_file_name,
                        record.document.page_number,
                        evidence.source_label,
                        evidence.standard_path,
                        self._excel_value(evidence.raw_value),
                        self._excel_value(evidence.normalized_value),
                        evidence.parser,
                        evidence.extraction_method.value,
                        evidence.transform,
                        evidence.confidence,
                        evidence.user_modified,
                    ]
                )
        for cell in sheet["K"][1:]:
            cell.number_format = "0%"

    @staticmethod
    def _transactions(records: list[ReceiptRecord]):
        for record in records:
            for transaction in record.transactions:
                yield record, transaction

    @staticmethod
    def _detail_keys(selected_columns: list[str] | None) -> list[str]:
        if selected_columns is None:
            return list(DETAIL_COLUMNS)
        if not selected_columns:
            return DEFAULT_DETAIL_KEYS.copy()
        keys = ["transaction_id"]
        keys.extend(key for key in selected_columns if key in DETAIL_COLUMNS and key != "transaction_id")
        return list(dict.fromkeys(keys))

    @staticmethod
    def _excel_value(value: Any) -> Any:
        if value is None or isinstance(value, (str, int, float, bool, datetime, date)):
            return value
        return str(value)

    @staticmethod
    def _status_label(status: ValidationStatus) -> str:
        return {
            ValidationStatus.NORMAL: "정상",
            ValidationStatus.REVIEW_REQUIRED: "확인 필요",
            ValidationStatus.ANALYSIS_FAILED: "분석 실패",
        }[status]

    @staticmethod
    def _status_conditional_format(sheet, column_letter: str, last_row: int) -> None:
        target = f"{column_letter}2:{column_letter}{last_row}"
        sheet.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'${column_letter}2="정상"'], fill=OK_FILL),
        )
        sheet.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'${column_letter}2="확인 필요"'], fill=WARNING_FILL),
        )
        sheet.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'${column_letter}2="분석 실패"'], fill=ERROR_FILL),
        )

    @staticmethod
    def _finish_sheet(sheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 26
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = SUBTLE_BORDER
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            maximum = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[: min(len(column_cells), 200)]
            )
            sheet.column_dimensions[letter].width = min(max(maximum + 2, 10), 38)
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{max(1, sheet.max_row)}"

    @staticmethod
    def _verify_reopen(path: Path) -> None:
        workbook = load_workbook(path, read_only=False, data_only=False)
        expected = ["세무정리", "상세내역", "오류내역", "원본매핑"]
        if workbook.sheetnames != expected:
            workbook.close()
            raise ValueError("Excel 시트 구성이 올바르지 않습니다.")
        for sheet in workbook.worksheets:
            if sheet.max_row < 1 or sheet.freeze_panes != "A2":
                workbook.close()
                raise ValueError(f"Excel 시트 검증 실패: {sheet.title}")
        workbook.close()

