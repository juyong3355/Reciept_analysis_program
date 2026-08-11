from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from receipt_mvp.exporters import ExcelExporter
from receipt_mvp.models import (
    Amounts,
    DocumentInfo,
    ExtractionMethod,
    FieldEvidence,
    LineItem,
    PageExtraction,
    Party,
    Platform,
    ReceiptRecord,
    Transaction,
    ValidationStatus,
)


def make_record() -> ReceiptRecord:
    transaction = Transaction(
        occurred_at=datetime(2026, 1, 2, 3, 4, 5),
        order_number="00123",
        approval_number="00042",
        evidence_type="신용카드 매출전표",
        items=[LineItem(name="상품 A"), LineItem(name="상품 B")],
        seller=Party(name="합성판매자", business_registration_number="1234567890"),
        merchant=Party(name="합성가맹점", business_registration_number="1112233333"),
        amounts=Amounts(
            supply_amount=10000,
            tax_exempt_amount=0,
            vat_amount=1000,
            service_charge=0,
            approved_amount=11000,
            total_amount=11000,
        ),
        confidence=0.97,
        validation_status=ValidationStatus.NORMAL,
    )
    return ReceiptRecord(
        document=DocumentInfo(
            source_file_name="synthetic.pdf",
            source_file_path="C:/synthetic.pdf",
            file_sha256="a" * 64,
            page_number=1,
            platform=Platform.NAVER,
            document_type="CARD_RECEIPT",
            extraction_method=ExtractionMethod.OCR,
        ),
        extraction=PageExtraction(
            source_path="C:/synthetic.pdf", page_number=1, method=ExtractionMethod.OCR
        ),
        transactions=[transaction],
        field_evidence=[
            FieldEvidence(
                transaction_id=transaction.transaction_id,
                source_label="합계",
                standard_path="amounts.total_amount",
                raw_value="11,000",
                normalized_value=11000,
                parser="SyntheticParser",
                extraction_method=ExtractionMethod.OCR,
                confidence=0.97,
            )
        ],
    )


def test_export_creates_four_reopenable_sheets_with_types(tmp_path: Path) -> None:
    path = ExcelExporter().export([make_record()], tmp_path / "result.xlsx")
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["세무정리", "상세내역", "오류내역", "원본매핑"]
    tax = workbook["세무정리"]
    assert tax["A2"].data_type == "d"
    assert tax["D2"].data_type == "n"
    assert tax.freeze_panes == "A2"
    details = workbook["상세내역"]
    headers = [cell.value for cell in details[1]]
    assert "거래 ID" in headers
    assert details.max_row == 3
    workbook.close()


def test_custom_columns_keep_transaction_relation(tmp_path: Path) -> None:
    path = ExcelExporter().export(
        [make_record()], tmp_path / "custom.xlsx", ["file", "item_name", "total_amount"]
    )
    workbook = load_workbook(path)
    headers = [cell.value for cell in workbook["상세내역"][1]]
    assert headers == ["거래 ID", "원본파일", "상품명", "합계금액"]
    workbook.close()

